from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from selenium.webdriver.firefox.options import Options
from PIL import Image
from fpdf import FPDF


# variable
image_folder = "images/"
solve_me = []
solve_counter = 1

# code.
options = Options()
options.headless = True

driver = webdriver.Firefox(options=options)

wb = load_workbook(filename="isloreren.xlsx", read_only=True, data_only=True)
current_sheet = wb['Blad1']

for row in current_sheet.iter_rows(min_row=1, max_row=150, max_col=3, values_only=True):
    if row[2] is not None:
        spl = row[2].split("}{")
        if len(spl) == 2:
            p1 = f"{spl[0]}}}"
            spl2 = spl[1].split("=")
            p2 = f'{{{spl[1].split("}=")[0]}}}'
            p3 = f'={spl[1].split("}=")[1]}'
            solve_me.append([p1, p2, p3])

for s in solve_me:
    image_list = []
    print(f"Solving: {s}")

    driver.get("https://www.mathpapa.com/algebra-calculator.html")
    driver.set_window_size(1024, 8000)
    element = driver.switch_to.active_element

    element.send_keys(s[0])
    element.send_keys(Keys.DOWN)
    element.send_keys(s[1])
    element.send_keys(Keys.RIGHT)
    element.send_keys(s[2])

    element = driver.find_element(By.ID, "parse_btn")
    element.click()

    # save first
    calc = driver.find_element(By.ID, "source3")
    calc.screenshot(f"{image_folder}sum.png")
    image_list.append("sum.png")

    # scroll to location
    driver.execute_script("window.scrollTo(0, 380)")

    # check for multiple variable
    element = driver.find_element(By.ID, "solvarval")
    options = [x for x in element.find_elements(By.TAG_NAME, "option")]

    # walk through all variable
    for e in options:
        n = e.get_attribute("value")
        e.click()

        solve = driver.find_element(By.ID, "solvarout3")
        # take the screenshots:
        solve.screenshot(f"{image_folder}solve_{n}.png")
        image_list.append(f"solve_{n}.png")

    # create a PDF
    pdf = FPDF('P', 'pt', 'A4')
    pdf.add_page()

    margin = 20
    factor = 2
    itter = 0

    img_count = 0
    print(f"Images: {len(image_list)}")
    for image in image_list:
        img = Image.open(image_folder + image)

        width = img.width
        height = img.height

        if len(image_list) > 3 and img_count > 2:
            print("* Adding extra page")
            pdf.add_page()
            img_count = 1
            margin = 20

        print(f"* Adding image: {img_count}")
        pdf.image(image_folder + image, x=None, y=margin, w=int(width / factor), h=int(height / factor))

        # add a nice line
        if itter >= 1:
            pdf.line(10, margin, 600, margin)
        itter += 1

        margin += int(height / factor) + 10

        img_count += 1

    # write to disk
    pdf.output(f'output/uitwerking_{solve_counter}.pdf', 'F')
    print(f"Created PDF: uitwerking_{solve_counter}.pdf")

    solve_counter += 1
